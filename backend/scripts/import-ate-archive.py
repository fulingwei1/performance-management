#!/usr/bin/env python3
"""
Import ATE 人事档案系统.xlsx into the performance system.
Reads Excel, maps columns to employee fields, and calls backend API.
Usage:
  pip install openpyxl requests  # if needed
  API_URL=http://localhost:3001 python backend/scripts/import-ate-archive.py /path/to/ATE-人事档案系统.xlsx
"""
import os
import re
import sys

try:
    import openpyxl
    import requests
except ImportError as e:
    print("Please install: pip install openpyxl requests")
    sys.exit(1)

API_BASE = os.environ.get("API_URL", "http://localhost:3001")
API = f"{API_BASE}/api"
AUTH = {"username": "林作倩", "password": "123456", "role": "hr"}

# Map 级别 to system level
LEVEL_MAP = {
    "senior": re.compile(r"高级"),
    "intermediate": re.compile(r"中级|主管|专员|待定"),
    "junior": re.compile(r"初级"),
    "assistant": re.compile(r"助理"),
}


def map_level(raw: str) -> str:
    if not raw or not str(raw).strip():
        return "intermediate"
    raw = str(raw).strip()
    if LEVEL_MAP["senior"].search(raw):
        return "senior"
    if LEVEL_MAP["assistant"].search(raw):
        return "assistant"
    if LEVEL_MAP["junior"].search(raw):
        return "junior"
    if LEVEL_MAP["intermediate"].search(raw):
        return "intermediate"
    return "intermediate"


def norm(s):
    if s is None:
        return ""
    t = str(s).strip().replace("\n", " ").strip()
    return t if t and t != "/" else ""


def build_org_tree_from_rows(rows):
    """Build unique (一级, 二级, 三级) hierarchy from Excel rows. Returns (roots, level2_set, level3_set)."""
    roots = set()  # 一级
    level2_set = set()  # (一级, 二级)
    level3_set = set()  # (一级, 二级, 三级)
    for row in rows[1:]:
        if not norm(row[0]):
            continue
        d1 = norm(row[2]) or "未分配"
        d2 = norm(row[3])
        d3 = norm(row[4])
        roots.add(d1)
        if d2:
            level2_set.add((d1, d2))
        if d2 and d3:
            level3_set.add((d1, d2, d3))
    return sorted(roots), sorted(level2_set), sorted(level3_set)


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else "/Users/flw/Desktop/ATE-人事档案系统.xlsx"
    if not os.path.isfile(path):
        print(f"File not found: {path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "员工信息表" not in wb.sheetnames:
        print("Sheet '员工信息表' not found. Available:", wb.sheetnames)
        sys.exit(1)
    ws = wb["员工信息表"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        print("No data rows")
        sys.exit(0)

    # Build org tree from 一级/二级/三级
    roots, level2_set, level3_set = build_org_tree_from_rows(rows)
    print(f"Org tree: {len(roots)} roots, {len(level2_set)} level2, {len(level3_set)} level3.")

    # Header: 姓名, 在职离职状态, 一级部门, 二级部门, 三级部门, 直接上级, 岗位, 级别, 性别, 是否转正, 身份证号
    employees = []
    for i, row in enumerate(rows[1:], start=2):
        name = norm(row[0])
        if not name:
            continue
        dept1 = norm(row[2]) or "未分配"
        dept2 = norm(row[3])
        dept3 = norm(row[4])
        sub_dept = dept3 if dept3 else (dept2 if dept2 else dept1)
        level_raw = norm(row[7])
        emp_id = f"ate-{i-1}"
        employees.append({
            "id": emp_id,
            "name": name,
            "department": dept1,
            "subDepartment": sub_dept,
            "role": "employee",
            "level": map_level(level_raw),
            "managerId": None,
            "password": "123456",
        })

    print(f"Parsed {len(employees)} employees from Excel.")

    # Login
    r = requests.post(f"{API_BASE}/api/auth/login", json=AUTH, timeout=10)
    if r.status_code != 200 or not r.json().get("success"):
        print("Login failed:", r.status_code, r.text[:200])
        sys.exit(1)
    token = r.json().get("data", {}).get("token") or r.json().get("token")
    if not token:
        print("No token in response:", r.json())
        sys.exit(1)
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    # Fetch existing departments to avoid duplicates (code -> id)
    code_to_id = {}
    try:
        r = requests.get(f"{API}/organization/departments", headers=headers, timeout=10)
        if r.status_code == 200 and r.json().get("success"):
            for d in r.json().get("data") or []:
                code_to_id[d.get("code") or ""] = d.get("id")
    except Exception as e:
        print("Warning: could not load existing departments:", e)

    # Create organization tree: roots (一级) -> level2 (二级) -> level3 (三级)
    def create_dept(name, code, parent_id, sort_order):
        if code in code_to_id and code_to_id[code]:
            return code_to_id[code]
        payload = {"name": name, "code": code, "sortOrder": sort_order}
        if parent_id:
            payload["parentId"] = parent_id
        resp = requests.post(f"{API}/organization/departments", json=payload, headers=headers, timeout=10)
        if resp.status_code in (200, 201) and resp.json().get("success"):
            new_id = resp.json().get("data", {}).get("id")
            if new_id:
                code_to_id[code] = new_id
                return new_id
        return None

    def refresh_code_to_id():
        r = requests.get(f"{API}/organization/departments", headers=headers, timeout=10)
        if r.status_code == 200 and r.json().get("success"):
            for x in r.json().get("data") or []:
                code_to_id[x.get("code") or ""] = x.get("id")

    for idx, d1 in enumerate(roots):
        create_dept(d1, d1, None, idx)
    refresh_code_to_id()
    for (d1, d2) in level2_set:
        code2 = f"{d1}-{d2}"
        parent_id = code_to_id.get(d1)
        if parent_id:
            create_dept(d2, code2, parent_id, len(code_to_id))
    refresh_code_to_id()
    for (d1, d2, d3) in level3_set:
        code2 = f"{d1}-{d2}"
        code3 = f"{d1}-{d2}-{d3}"
        parent_id = code_to_id.get(code2)
        if parent_id:
            create_dept(d3, code3, parent_id, len(code_to_id))

    print(f"Organization tree: ensured {len(code_to_id)} department nodes.")

    success = 0
    fail = 0
    for emp in employees:
        try:
            resp = requests.post(f"{API}/employees", json=emp, headers=headers, timeout=10)
            if resp.status_code in (200, 201) and (resp.json() or {}).get("success"):
                success += 1
            else:
                fail += 1
                msg = (resp.json() or {}).get("error") or resp.text[:80]
                print(f"  Fail {emp['id']} {emp['name']}: {msg}")
        except Exception as e:
            fail += 1
            print(f"  Error {emp['id']} {emp['name']}: {e}")

    print(f"Import done: success={success}, fail={fail}.")


if __name__ == "__main__":
    main()
